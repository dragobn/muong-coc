#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Dựng file Excel cung cấp thông tin 17 tour/cung Mường Cốc (9 tour + 8 cung đạp xe).
Nguồn: data/tours/*.json (bản chốt GĐ20). Hotline chính thức Ban điều phối.
Cột theo mẫu đối tác: Thời lượng | Chi phí | Hotline | Link ảnh | Ghi chú | Tóm tắt nội dung.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.join(os.path.dirname(__file__), "..", "data", "tours")
OUT = os.path.join(os.path.dirname(__file__), "..", "..", "..", "..",
                   "Thông tin các tour Mường Cốc - 260616.xlsx")

HOTLINE = "0962 029 198"
EMAIL = "dulichcongdongmuongcoc.myduc@gmail.com"
DON_VI = "Ban Điều phối Du lịch cộng đồng Mường Cốc"
SITE = "https://dulichcongdongmuongcoc.vn/cung/"  # link trang chương trình (có gallery ảnh)

# Giá ĐỀ XUẤT cho 9 tour (đ/khách, nhóm 6–10 khách) — neo theo cung đạp xe 650k/ngày.
# Cung C1–C8 giữ 650k (lấy trực tiếp từ data). Anh điều chỉnh sau.
PRICE = {
    "tour-1n-thien-nhien": "690.000 đ/khách (đề xuất)",
    "tour-1n-van-hoa-sen": "790.000 đ/khách (đề xuất)",
    "tour-2n-thien-nhien": "1.750.000 đ/khách (đề xuất)",
    "tour-2n-van-hoa-sen": "1.850.000 đ/khách (đề xuất)",
    "tour-3n-huong-tich": "2.950.000 đ/khách (đề xuất)",
    "tour-kn-quang-phu-cau": "850.000 đ/khách (đề xuất)",
    "tour-kn-chua-huong": "1.150.000 đ/khách (đề xuất)",
    "tour-kn-cuc-phuong": "2.450.000 đ/khách (đề xuất)",
    "tour-kn-mai-chau-pu-luong": "3.950.000 đ/khách (đề xuất, 3N2Đ; 4N +900.000)",
}

# Cơ sở tính giá đề xuất (để anh điều chỉnh minh bạch)
PRICE_BASIS = [
    ("Tour 1 – Thiên nhiên 1N", "690.000", "Bằng cung đạp xe (650k) + giao lưu văn nghệ cộng đồng + trồng cây."),
    ("Tour 2 – Sen & Mo Mường 1N", "790.000", "Base 1N + workshop trà sen + trải nghiệm đầm sen + quà cộng đồng. (Lễ Mo Mường phụ thu riêng.)"),
    ("Tour 3 – Thiên nhiên 2N1Đ", "1.750.000", "≈ ngày 1 (650k) + 1 đêm homestay (~250k) + tối+sáng (~260k) + hoạt động ngày 2 (~450k) + lửa trại (~120k)."),
    ("Tour 4 – Sen & Mo Mường 2N1Đ", "1.850.000", "Như tour 3 + workshop sen/quà; phần trải nghiệm văn hoá đậm hơn. (Mo Mường phụ thu.)"),
    ("Tour 5 – Trekking Hương Tích 3N2Đ", "2.950.000", "Base 1N + 2 đêm homestay & ăn (~1.020k) + trek có người dẫn (~500k) + vé Hương Tích (~600k) + hướng dẫn trek (~180k)."),
    ("Tour 6 – Quảng Phú Cầu 1N", "850.000", "Base 1N + xe đưa đón liên vùng (làng hương ~Ứng Hoà) cao hơn cung nội vùng (~200k)."),
    ("Tour 7 – Chùa Hương 1N", "1.150.000", "Base 1N + vé thuyền suối Yến + vé quần thể Chùa Hương + di chuyển liên vùng (~300–350k vé/phí)."),
    ("Tour 8 – Cúc Phương 2N1Đ", "2.450.000", "≈ ngày 1 (650k) + 1 đêm homestay & ăn (~510k) + xe liên tỉnh Ninh Bình + vé VQG Cúc Phương & cứu hộ (~700k) + ngày 2 (~600k)."),
    ("Tour 9 – Mai Châu–Pù Luông 3–4N", "3.950.000", "Hành trình liên tỉnh dài nhất: 1 đêm MC + 1 đêm Mai Châu/Pù Luông, xe đường dài, vé trải nghiệm. Giá cho 3N2Đ; bản 4N cộng ~900k."),
    ("Cung đạp xe C1–C8 (1N)", "650.000", "Giá đã chốt trong data: gồm trưa Cỗ Lá Mường, HDV bản địa, xe đạp+mũ, vé+hoạt động, bảo hiểm. Nhóm 2–12 khách."),
]

# (mã, slug, nhãn nhóm)
TOURS = [
    ("Tour 1", "tour-1n-thien-nhien", "Tour nội vùng"),
    ("Tour 2", "tour-1n-van-hoa-sen", "Tour nội vùng"),
    ("Tour 3", "tour-2n-thien-nhien", "Tour nội vùng"),
    ("Tour 4", "tour-2n-van-hoa-sen", "Tour nội vùng"),
    ("Tour 5", "tour-3n-huong-tich", "Tour nội vùng"),
    ("Tour 6", "tour-kn-quang-phu-cau", "Tour kết nối"),
    ("Tour 7", "tour-kn-chua-huong", "Tour kết nối"),
    ("Tour 8", "tour-kn-cuc-phuong", "Tour kết nối"),
    ("Tour 9", "tour-kn-mai-chau-pu-luong", "Tour kết nối"),
    ("Cung C1", "vong-xe-xanh", "Cung đạp xe"),
    ("Cung C2", "theo-dong-ai-nang", "Cung đạp xe"),
    ("Cung C3", "dau-xua-muong-coc", "Cung đạp xe"),
    ("Cung C4", "hon-muong-ho-baiboo", "Cung đạp xe"),
    ("Cung C5", "ve-mien-di-san", "Cung đạp xe"),
    ("Cung C6", "xanh-giua-long", "Cung đạp xe"),
    ("Cung C7", "huong-dong-bo-moi", "Cung đạp xe"),
    ("Cung C8", "nhip-song-bo-moi", "Cung đạp xe"),
]

def load(slug):
    with open(os.path.join(BASE, slug + ".json"), encoding="utf-8") as f:
        return json.load(f)

def tom_tat(d):
    """Tóm tắt nội dung: tagline + intro gọn + lịch trình rút gọn."""
    parts = []
    if d.get("tagline"):
        parts.append(d["tagline"].strip().rstrip(".") + ".")
    intro = (d.get("intro") or "").strip()
    if intro:
        # lấy ~2 câu đầu cho gọn
        cau = intro.replace("\n", " ").split(". ")
        parts.append(". ".join(cau[:2]).strip().rstrip(".") + ".")
    lt = d.get("lich_trinh") or []
    if lt:
        chang = []
        for x in lt:
            g = (x.get("gio_hoac_buoi") or "").strip()
            diem = (x.get("diem") or "").strip()
            if diem:
                chang.append(f"{g}: {diem}" if g else diem)
        if chang:
            parts.append("Lịch trình: " + " → ".join(chang) + ".")
    return "\n".join(parts)

def ghi_chu(d, ma):
    rows = []
    if d.get("quy_mo"):
        rows.append(f"Quy mô: {d['quy_mo']}.")
    rows.append("Xuất phát & kết thúc tại Tourism Hub Đồi Dùng, Mường Cốc.")
    # highlights tối đa 3
    hl = d.get("highlights") or []
    if hl:
        rows.append("Điểm nhấn: " + " ".join("• " + h.strip().rstrip(".") + "." for h in hl[:3]))
    # lưu ý đặt trước
    for x in (d.get("luu_y") or []):
        td = (x.get("tieu_de") or "").strip()
        if "mo mường" in td.lower() or "đặt" in (x.get("mo_ta") or "").lower():
            rows.append(f"Lưu ý: {td} cần đặt trước với Ban điều phối.")
            break
    if ma.startswith("Cung"):
        rows.append("Giá đã gồm bữa trưa Cỗ Lá Mường.")
    return "\n".join(rows)

# ---- build workbook ----
wb = Workbook()
ws = wb.active
ws.title = "Tour Mường Cốc"

headers = ["STT", "Mã", "Tên chương trình", "Loại hình",
           "Thời lượng tour\n(từ khi bắt đầu đến khi kết thúc tour)",
           "Chi phí tour", "Hotline liên hệ đặt tour",
           "Link hình ảnh kèm theo từng chương trình",
           "Ghi chú", "Tóm tắt nội dung"]

green = PatternFill("solid", fgColor="2E6B4F")
white_bold = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
body_font = Font(name="Times New Roman", size=12)
thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")

# header row
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = green; cell.font = white_bold; cell.alignment = center; cell.border = border
ws.row_dimensions[1].height = 46

r = 2
for stt, (ma, slug, nhom) in enumerate(TOURS, 1):
    d = load(slug)
    vals = [
        stt, ma, d.get("ten", slug), nhom,
        d.get("thoi_luong", ""),
        PRICE.get(slug, d.get("gia", "")),
        f"{HOTLINE}\n{EMAIL}",
        f"{SITE}{slug}/",  # link trang chương trình (gallery ảnh)
        ghi_chu(d, ma),
        tom_tat(d),
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = body_font; cell.border = border
        cell.alignment = center if c in (1, 2, 4) else wrap_top
    r += 1

# widths
widths = {1: 5, 2: 8, 3: 30, 4: 14, 5: 20, 6: 22, 7: 24, 8: 22, 9: 42, 10: 60}
from openpyxl.utils import get_column_letter
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A2"

# sheet ghi chú nguồn
ws2 = wb.create_sheet("Thông tin liên hệ")
info = [
    ["Đơn vị vận hành", DON_VI],
    ["Hotline đặt tour", HOTLINE],
    ["Email", EMAIL],
    ["Điểm xuất phát chung", "Tourism Hub Đồi Dùng, Mường Cốc (xã Mỹ Đức, Hà Nội)"],
    ["Số chương trình", "17 (9 tour + 8 cung đạp xe)"],
    ["Giá cung đạp xe", "650.000 đ/khách/ngày (gồm bữa trưa Cỗ Lá Mường), nhóm 2–12 khách"],
    ["Giá tour (1–9)", "Số trong cột Chi phí là GIÁ ĐỀ XUẤT (nhóm 6–10 khách) — xem sheet 'Cơ sở tính giá'. Nhóm 2–4 khách phụ thu 15–25%."],
    ["Link hình ảnh", "Trỏ tới trang chương trình live (có gallery ảnh) tại dulichcongdongmuongcoc.vn. Album ảnh rời để trim nằm ở thư mục demo-site/album-anh/."],
    ["Website", "https://dulichcongdongmuongcoc.vn"],
]
for i, (k, v) in enumerate(info, 1):
    a = ws2.cell(row=i, column=1, value=k); a.font = Font(name="Times New Roman", size=12, bold=True)
    b = ws2.cell(row=i, column=2, value=v); b.font = body_font; b.alignment = Alignment(wrap_text=True, vertical="top")
ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 78

# sheet cơ sở tính giá
ws3 = wb.create_sheet("Cơ sở tính giá")
ph = ["Chương trình", "Giá đề xuất (đ/khách)", "Cơ sở tính"]
for c, h in enumerate(ph, 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.fill = green; cell.font = white_bold; cell.alignment = center; cell.border = border
for i, (ten, gia, co_so) in enumerate(PRICE_BASIS, 2):
    a = ws3.cell(row=i, column=1, value=ten)
    a.font = Font(name="Times New Roman", size=12, bold=True)
    a.alignment = Alignment(wrap_text=True, vertical="top"); a.border = border
    b = ws3.cell(row=i, column=2, value=gia)
    b.font = body_font; b.alignment = Alignment(horizontal="center", vertical="top"); b.border = border
    cs = ws3.cell(row=i, column=3, value=co_so)
    cs.font = body_font; cs.alignment = Alignment(wrap_text=True, vertical="top"); cs.border = border
ws3.cell(row=1, column=1).border = border
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 80
note = ws3.cell(row=len(PRICE_BASIS)+3, column=1,
                value="Lưu ý: giá đề xuất nội bộ, chưa gồm VAT, chưa gồm đưa đón Hà Nội ↔ Mường Cốc (nếu khách yêu cầu). Anh điều chỉnh trước khi gửi đối tác.")
note.font = Font(name="Times New Roman", size=11, italic=True)

wb.save(OUT)
print("SAVED:", os.path.abspath(OUT))
print("Rows:", len(TOURS))
