#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build KML mới theo cấu trúc 2-tầng cho Google My Maps:
  A) demo-site/kml/diem-den/{loai-slug}.kml  — điểm đến gom theo PHÂN LOẠI (cột Loại của Excel)
  B) demo-site/kml/tuyen/{slug}.kml          — chỉ LineString tách từ KML cũ (không marker)

Nguồn:
  - Database MC/muongcoc-poi-thu-thap-hop-260512.xlsx (sheet "POI Mường Cốc") — phân loại CHÍNH
  - Database MC/muongcoc-cbt-database-v2-260511-0150.xlsx (02_supplier, 06_asset_story) — mô tả + SĐT bổ sung
  - data/master-poi-coords.json — toạ độ EXIF
  - kml/{slug}.kml (cũ) — tách LineString
KHÔNG bịa: chỉ điền field có dữ liệu thật.
"""
import json, re, unicodedata, html, os
import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
PROJ = os.path.abspath(os.path.join(ROOT, "..", "..", ".."))
DB = os.path.join(PROJ, "Database MC")
XLSX_POI = os.path.join(DB, "muongcoc-poi-thu-thap-hop-260512.xlsx")
XLSX_V2 = os.path.join(DB, "muongcoc-cbt-database-v2-260511-0150.xlsx")
COORDS = os.path.join(ROOT, "data", "master-poi-coords.json")
KML_OLD = os.path.join(ROOT, "kml")
OUT_DD = os.path.join(KML_OLD, "diem-den")
OUT_TU = os.path.join(KML_OLD, "tuyen")

# anchor toạ độ trung tâm các thôn (fallback cho điểm thiếu GPS) — lấy từ coords cùng thôn
THON_FALLBACK = {}  # filled later from matched points

# ---- mapping Loại -> layer phân loại (slug, tên hiển thị, màu icon KML aabbggrr, icon href) ----
# màu KML = aabbggrr (alpha,blue,green,red)
LAYER = {
    "canh-quan-thien-nhien": ("Cảnh quan thiên nhiên", "ff00aa00",
        "http://maps.google.com/mapfiles/kml/paddle/grn-blank.png"),
    "di-tich-tam-linh": ("Di tích & tâm linh", "ff0066cc",
        "http://maps.google.com/mapfiles/kml/paddle/red-blank.png"),
    "luu-tru": ("Lưu trú (homestay / farmstay)", "ff3399ff",
        "http://maps.google.com/mapfiles/kml/paddle/orange-blank.png"),
    "am-thuc": ("Ẩm thực & điểm dừng chân", "ff0099ff",
        "http://maps.google.com/mapfiles/kml/paddle/ylw-blank.png"),
    "trai-nghiem-nghe": ("Trải nghiệm & nghề truyền thống", "ffcc66ff",
        "http://maps.google.com/mapfiles/kml/paddle/pink-blank.png"),
    "di-san-phi-vat-the": ("Di sản phi vật thể (Mo Mường, nghề, lễ)", "ffff66cc",
        "http://maps.google.com/mapfiles/kml/paddle/purple-blank.png"),
}

LOAI_TO_LAYER = {
    "Hồ": "canh-quan-thien-nhien",
    "Thung": "canh-quan-thien-nhien",
    "Hang": "canh-quan-thien-nhien",
    "Rừng": "canh-quan-thien-nhien",
    "Đồi dược liệu": "canh-quan-thien-nhien",
    "Đình": "di-tich-tam-linh",
    "Chùa": "di-tich-tam-linh",
    "Đền / miếu": "di-tich-tam-linh",
    "Nhà thờ": "di-tich-tam-linh",
    "Nhà cổ": "di-tich-tam-linh",
    "Nhà văn hoá": "di-tich-tam-linh",
    "Homestay": "luu-tru",
    "Farmstay": "luu-tru",
    "Nhà hàng / F&B": "am-thuc",
    "Chợ": "am-thuc",
    "Nông trại trải nghiệm": "trai-nghiem-nghe",
    "Nghệ nhân văn hoá": "trai-nghiem-nghe",
    "Nghệ nhân nấu rượu": "trai-nghiem-nghe",
    "Văn hoá phi vật thể (ICH)": "di-san-phi-vat-the",
}

def strip_accent(s):
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D")

def norm(s):
    if not s: return ""
    s = strip_accent(str(s)).lower()
    s = re.sub(r"\([^)]*\)", " ", s)            # bỏ ngoặc
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def slugify(s):
    s = strip_accent(str(s)).lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:60]

# ---- load coords ----
coords_raw = json.load(open(COORDS, encoding="utf-8"))
coords_idx = {}  # norm(name) -> (lat,lon,anchor)
anchor_coord = {}  # anchor -> (lat,lon)
for name, v in coords_raw.items():
    lat, lon, anc = v.get("lat"), v.get("lon"), v.get("anchor")
    if lat is None or lon is None:
        continue
    coords_idx[norm(name)] = (lat, lon, anc)
    if anc and anc not in anchor_coord:
        anchor_coord[anc] = (lat, lon)

# token-based keyword anchors for fuzzy match
def find_coord(name):
    n = norm(name)
    if n in coords_idx:
        return coords_idx[n] + ("exact",)
    # token overlap: best jaccard, tie-break bằng việc khớp từ đầu (loại điểm)
    nt = set(n.split())
    first = n.split()[0] if n.split() else ""
    best, bestsc = None, 0.0
    for cn, val in coords_idx.items():
        ct = set(cn.split())
        if not ct: continue
        inter = nt & ct
        if not inter: continue
        sc = len(inter) / len(nt | ct)
        if first and first in ct:
            sc += 0.15  # ưu tiên khớp loại (đình/chùa/đền…) tránh nhầm Đình↔Chùa cùng "Phú Cốc"
        if sc > bestsc:
            bestsc, best = sc, val
    if best and bestsc >= 0.34:
        return best + ("fuzzy",)
    return None

# ---- load supplier (phone + name + gps) from v2 ----
wb2 = openpyxl.load_workbook(XLSX_V2, data_only=True)
sup_rows = list(wb2["02_supplier"].iter_rows(values_only=True))
sup_phone = {}  # norm(name) -> phone
sup_gps = {}    # norm(name) -> (lat,lon)
for r in sup_rows[3:]:
    if r and r[1]:
        nn = norm(r[1])
        if r[7]:
            sup_phone[nn] = str(r[7])
        if r[5] and r[6]:
            try:
                sup_gps[nn] = (float(r[5]), float(r[6]))
            except (TypeError, ValueError):
                pass

# asset_story narratives
story_rows = list(wb2["06_asset_story"].iter_rows(values_only=True))
story_map = {}  # asset-id slug-ish -> narrative
for r in story_rows[3:]:
    if r and r[0] and r[3]:
        story_map[norm(r[1] or "")] = str(r[3])

def find_story(name):
    # story chỉ gán khi tên khớp chặt (tránh nhầm Đình Phú Cốc ↔ Chùa Phú Cốc)
    n = norm(name)
    nt = set(n.split())
    nfirst = n.split()[0] if n.split() else ""
    best, bestsc = None, 0.0
    for k, v in story_map.items():
        if not k: continue
        kt = set(k.split())
        if not kt: continue
        kfirst = k.split()[0]
        if nfirst and kfirst and nfirst != kfirst:
            continue  # loại điểm khác nhau -> bỏ
        sc = len(nt & kt) / len(nt | kt)
        if sc > bestsc:
            bestsc, best = sc, v
    return best if bestsc >= 0.5 else None

def _best_token(name, chu, table, thr=0.4):
    for src in (name, chu):
        if not src: continue
        n = norm(src)
        if n in table:
            return table[n]
        nt = set(n.split())
        best, bestsc = None, 0.0
        for k, v in table.items():
            kt = set(k.split())
            if not kt: continue
            sc = len(nt & kt) / len(nt | kt)
            if sc > bestsc:
                bestsc, best = sc, v
        if best and bestsc >= thr:
            return best
    return None

def find_phone(name, chu):
    return _best_token(name, chu, sup_phone, 0.4)

def find_sup_gps(name, chu):
    return _best_token(name, chu, sup_gps, 0.45)

# ---- read POI Excel ----
wb = openpyxl.load_workbook(XLSX_POI, data_only=True)
ws = wb["POI Mường Cốc"]
rows = list(ws.iter_rows(values_only=True))

pois = []  # dict per POI
missing_gps = []
for r in rows[1:]:
    if not r or not r[1]:
        continue
    stt, ten, thon, loai, chu, sdt, brand, tt, ghi = (list(r) + [None]*9)[:9]
    layer = LOAI_TO_LAYER.get((loai or "").strip())
    if not layer:
        layer = "trai-nghiem-nghe"  # fallback
    c = find_coord(ten)
    lat = lon = None
    matchtype = None
    if c and c[3] == "exact":
        lat, lon, matchtype = c[0], c[1], "exact"
    else:
        # coords fuzzy không đáng tin bằng supplier GPS chính xác từng hộ
        sg = find_sup_gps(ten, chu)
        if sg:
            lat, lon, matchtype = sg[0], sg[1], "supplier-gps"
        elif c:
            lat, lon, matchtype = c[0], c[1], "fuzzy"
    pois.append(dict(stt=stt, ten=ten, thon=thon, loai=loai, chu=chu, sdt=sdt,
                     brand=brand, tt=tt, ghi=ghi, layer=layer, lat=lat, lon=lon,
                     match=matchtype))

# build thon fallback from matched coords (centroid per thon)
from collections import defaultdict
thon_pts = defaultdict(list)
for p in pois:
    if p["lat"] is not None and p["thon"]:
        thon_pts[norm(p["thon"])].append((p["lat"], p["lon"]))
thon_centroid = {k: (sum(a for a,b in v)/len(v), sum(b for a,b in v)/len(v)) for k,v in thon_pts.items()}
HUB = anchor_coord.get("hub", (20.6224767, 105.6897842))

for p in pois:
    if p["lat"] is None:
        # fallback cuối: tâm thôn (toạ độ tạm — cần kéo lại)
        ck = norm(p["thon"] or "")
        if ck in thon_centroid:
            p["lat"], p["lon"] = thon_centroid[ck]
            p["coord_flag"] = "TẠM (tâm thôn %s)" % p["thon"]
        else:
            p["lat"], p["lon"] = HUB
            p["coord_flag"] = "TẠM (Hub Đồi Dùng)"
        missing_gps.append(p)
    else:
        p["coord_flag"] = ""

# ---- build description HTML ----
def desc_html(p):
    parts = []
    story = find_story(p["ten"])
    if story:
        parts.append("<p>%s</p>" % html.escape(story))
    else:
        parts.append("<p>Điểm thuộc loại <b>%s</b>, thôn %s.</p>"
                     % (html.escape(str(p["loai"] or "")), html.escape(str(p["thon"] or ""))))
    # Dịch vụ block
    dv = []
    phone = p["sdt"] or find_phone(p["ten"], p["chu"])
    if p["chu"]:
        dv.append("Chủ/Phụ trách: %s" % html.escape(str(p["chu"])))
    if p["brand"] and str(p["brand"]).strip() not in ("(chưa có biển)","(chưa có biển)"):
        dv.append("Biển hiệu: %s" % html.escape(str(p["brand"])))
    if phone:
        dv.append("SĐT: %s" % html.escape(str(phone)))
    if p["tt"]:
        dv.append("Tình trạng: %s" % html.escape(str(p["tt"])))
    if p["coord_flag"]:
        dv.append("<i>Toạ độ %s — kéo lại đúng vị trí</i>" % html.escape(p["coord_flag"]))
    if dv:
        parts.append("<p><b>Dịch vụ:</b><br/>" + "<br/>".join(dv) + "</p>")
    return "<![CDATA[" + "".join(parts) + "]]>"

def kml_placemark(p):
    return (
        "  <Placemark>\n"
        "    <name>%s</name>\n"
        "    <styleUrl>#poi</styleUrl>\n"
        "    <description>%s</description>\n"
        "    <Point><coordinates>%s,%s,0</coordinates></Point>\n"
        "  </Placemark>\n"
    ) % (html.escape(str(p["ten"])), desc_html(p), p["lon"], p["lat"])

def write_layer_kml(layer_slug, name, color, icon, items):
    body = []
    body.append('<?xml version="1.0" encoding="UTF-8"?>\n')
    body.append('<kml xmlns="http://www.opengis.net/kml/2.2"><Document>\n')
    body.append("  <name>Mường Cốc — %s</name>\n" % html.escape(name))
    body.append('  <Style id="poi"><IconStyle><color>%s</color><scale>1.0</scale>'
                '<Icon><href>%s</href></Icon></IconStyle>'
                '<LabelStyle><scale>0.8</scale></LabelStyle></Style>\n' % (color, icon))
    for p in items:
        body.append(kml_placemark(p))
    body.append("</Document></kml>\n")
    open(os.path.join(OUT_DD, layer_slug + ".kml"), "w", encoding="utf-8").write("".join(body))

# group + write
layer_groups = defaultdict(list)
for p in pois:
    layer_groups[p["layer"]].append(p)

summary_A = []
for slug, (name, color, icon) in LAYER.items():
    items = layer_groups.get(slug, [])
    if not items: continue
    write_layer_kml(slug, name, color, icon, items)
    summary_A.append((slug, name, len(items)))

# ============ OUTPUT B: tách LineString ============
SLUG_KIND = {
    "vong-xe-xanh":"xe-dap","theo-dong-ai-nang":"xe-dap","dau-xua-muong-coc":"xe-dap",
    "hon-muong-ho-baiboo":"xe-dap","ve-mien-di-san":"xe-dap","xanh-giua-long":"xe-dap",
    "huong-dong-bo-moi":"xe-dap","nhip-song-bo-moi":"xe-dap",
    "tour-1n-thien-nhien":"tour","tour-1n-van-hoa-sen":"tour","tour-2n-thien-nhien":"tour",
    "tour-2n-van-hoa-sen":"tour","tour-3n-huong-tich":"tour","tour-kn-quang-phu-cau":"tour",
    "tour-kn-chua-huong":"tour","tour-kn-cuc-phuong":"tour","tour-kn-mai-chau-pu-luong":"tour",
    "trek-roc-eo":"trek","trek-xuyen-rung-huong-tich":"trek",
}
KIND_STYLE = {
    "xe-dap": ("ff00aa00", 5),   # xanh lá
    "tour":   ("ffcc6633", 5),   # cam xanh dương
    "trek":   ("ff3333cc", 5),   # đỏ
}
NAME_MAP = {
    "vong-xe-xanh":"Vòng Xe Xanh Mường Cốc","theo-dong-ai-nang":"Theo Dòng Ái Nàng",
    "dau-xua-muong-coc":"Dấu Xưa Mường Cốc","hon-muong-ho-baiboo":"Hồn Mường Bên Hồ Baibóo",
    "ve-mien-di-san":"Về Miền Di Sản","xanh-giua-long":"Xanh Giữa Lòng Mường Cốc",
    "huong-dong-bo-moi":"Hương Đồng Bơ Môi","nhip-song-bo-moi":"Nhịp Sống Mường Bơ Môi",
    "tour-1n-thien-nhien":"Tour 1 ngày — Thiên nhiên & nhà nông",
    "tour-1n-van-hoa-sen":"Tour 1 ngày — Văn hoá sen & Mo Mường",
    "tour-2n-thien-nhien":"Tour 2N1Đ — Phương án A","tour-2n-van-hoa-sen":"Tour 2N1Đ — Phương án B",
    "tour-3n-huong-tich":"Tour 3N2Đ — đến Hương Tích","tour-kn-quang-phu-cau":"Kết nối · Quảng Phú Cầu",
    "tour-kn-chua-huong":"Kết nối · Chùa Hương","tour-kn-cuc-phuong":"Kết nối · Cúc Phương",
    "tour-kn-mai-chau-pu-luong":"Kết nối · Mai Châu–Pù Luông",
    "trek-roc-eo":"Đường Mòn Rộc Éo","trek-xuyen-rung-huong-tich":"Xuyên Rừng Hương Tích",
}

summary_B = []
for fn in sorted(os.listdir(KML_OLD)):
    if not fn.endswith(".kml"): continue
    slug = fn[:-4]
    if slug not in SLUG_KIND: continue
    txt = open(os.path.join(KML_OLD, fn), encoding="utf-8").read()
    m = re.search(r"<LineString>.*?</LineString>", txt, re.S)
    if not m:
        continue
    line = m.group(0)
    kind = SLUG_KIND[slug]
    color, width = KIND_STYLE[kind]
    name = NAME_MAP.get(slug, slug)
    out = []
    out.append('<?xml version="1.0" encoding="UTF-8"?>\n')
    out.append('<kml xmlns="http://www.opengis.net/kml/2.2"><Document>\n')
    out.append("  <name>Tuyến — %s</name>\n" % html.escape(name))
    out.append('  <Style id="route"><LineStyle><color>%s</color><width>%d</width></LineStyle></Style>\n' % (color, width))
    out.append("  <Placemark>\n    <name>%s</name>\n    <styleUrl>#route</styleUrl>\n    %s\n  </Placemark>\n" % (html.escape(name), line))
    out.append("</Document></kml>\n")
    open(os.path.join(OUT_TU, slug + ".kml"), "w", encoding="utf-8").write("".join(out))
    summary_B.append((slug, kind))

# ---- report ----
print("=== OUTPUT A — LAYER ĐIỂM ĐẾN ===")
for slug, name, n in summary_A:
    print(f"  {slug:28} {name:42} {n} điểm")
print("=== OUTPUT B — TUYẾN ===")
xd = [s for s,k in summary_B if k=="xe-dap"]
print(f"  xe-dap: {len(xd)} | tour: {len([1 for _,k in summary_B if k=='tour'])} | trek: {len([1 for _,k in summary_B if k=='trek'])} | tổng {len(summary_B)}")
print("=== ĐIỂM THIẾU GPS (toạ độ tạm) ===")
for p in missing_gps:
    print(f"  [{p['stt']}] {p['ten']} (thôn {p['thon']}) -> {p['coord_flag']}")
print("=== ĐIỂM MATCH FUZZY (kiểm lại) ===")
for p in pois:
    if p["match"] == "fuzzy":
        print(f"  [{p['stt']}] {p['ten']} -> fuzzy {p['lat']},{p['lon']}")
